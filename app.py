"""
Flask — Test Plan Dashboard
Routes:
  GET  /        → dashboard UI
  GET  /data    → all sheets as JSON
  GET  /effort  → all sheets effort data
  POST /upload  → replace active Excel file
  GET  /ping    → keep-alive
"""

import re
from pathlib import Path
from flask import Flask, jsonify, render_template_string, request
import openpyxl

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20 MB

BASE        = Path(__file__).parent
DEFAULT_XL  = BASE / "Abbreviation of Test Plan_1P _ Better Meter.xlsx"
UPLOADED_XL = BASE / "uploaded_data.xlsx"

DISPLAY = {
    "Priority Test":                   "Priority Test",
    "Abbreviation of Test Plan_HW":    "HW Test Plan",
    "Abbreviation of Test Plan_FW":    "FW Test Plan",
    "Abbreviation of TP_Module HW":    "Module HW",
    "Abbreviation of TP_CommsTesting": "Comms Testing",
}

def active_xl():
    return UPLOADED_XL if UPLOADED_XL.exists() else DEFAULT_XL

def clean(v):
    return "" if v is None else str(v).replace("\n", " ").strip()

def to_float(v):
    if v is None: return None
    try: return float(v)
    except: return None

def parse_time_str(s):
    """Parse '15–20 min', '5 min' → average minutes as float."""
    if not s: return None
    s = re.sub(r'[–—-]', '-', str(s).lower()).replace('min','').replace('mins','').strip()
    if '-' in s:
        parts = s.split('-')
        try: return (float(parts[0].strip()) + float(parts[1].strip())) / 2
        except: return None
    try: return float(s.strip())
    except: return None

def find_hdr(rows):
    for i, r in enumerate(rows[:5]):
        if sum(1 for v in r if v is not None and str(v).strip()) >= 2:
            return i
    return 0

# ─── read_excel ───────────────────────────────────────────────────────────────
def read_excel():
    wb  = openpyxl.load_workbook(active_xl(), read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        ws   = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows: continue
        display = DISPLAY.get(name, name)
        if name == "Abbreviation of Test Plan_HW":
            hdr_rows = [rows[2], rows[3]]
            ncols = max(len(r) for r in hdr_rows)
            hdrs  = []
            for c in range(ncols):
                parts = []
                for r in hdr_rows:
                    v = clean(r[c]) if c < len(r) else ""
                    if v and v.lower() not in ("none","false","true"): parts.append(v)
                hdrs.append(" / ".join(parts) if parts else f"Col{c+1}")
            data = rows[4:]
        elif name in ("Abbreviation of TP_Module HW","Abbreviation of TP_CommsTesting"):
            hdrs = [clean(v) or f"Col{i+1}" for i,v in enumerate(rows[1])]
            data = rows[2:]
        else:
            hi   = find_hdr(rows)
            hdrs = [clean(v) or f"Col{i+1}" for i,v in enumerate(rows[hi])]
            data = rows[hi+1:]
        cleaned = [[clean(v) for v in r] for r in data if any(v is not None and str(v).strip() for v in r)]
        out[display] = {"headers": hdrs, "rows": cleaned}
    wb.close()
    return out

# ─── read_effort ──────────────────────────────────────────────────────────────
def read_effort():
    wb   = openpyxl.load_workbook(active_xl(), read_only=True, data_only=True)
    out  = {}

    # ── Priority Test ──────────────────────────────────────────────────────
    if "Priority Test" in wb.sheetnames:
        rows  = list(wb["Priority Test"].iter_rows(values_only=True))
        items = []
        for i, row in enumerate(rows[1:]):
            name = clean(row[1]) if len(row) > 1 else ""
            if not name: continue
            md = to_float(row[4]) if len(row) > 4 else None
            items.append({
                "id": f"pt_{i}", "name": name, "group": "Priority Tests",
                "days": md, "display": f"{md} days" if md is not None else "—",
                "level": None, "is_group": False,
            })
        out["Priority Test"] = {"unit": "days", "items": items}

    # ── HW Test Plan ───────────────────────────────────────────────────────
    if "Abbreviation of Test Plan_HW" in wb.sheetnames:
        rows  = list(wb["Abbreviation of Test Plan_HW"].iter_rows(values_only=True))
        items = []
        current_group = "General"
        for row in rows[4:]:
            sno  = row[0]
            name = clean(row[1]) if len(row) > 1 else ""
            if not name: continue
            level = None
            try: level = int(row[2]) if row[2] is not None else None
            except: pass
            if isinstance(sno, str) and sno.strip().isalpha():
                current_group = name
                items.append({"id": f"hw_grp_{sno}", "name": name, "group": name,
                              "days": None, "display": "—", "level": None, "is_group": True})
            elif isinstance(sno, (int, float)):
                md = to_float(row[23]) if len(row) > 23 else None
                items.append({"id": f"hw_{int(sno)}_{len(items)}", "name": name,
                              "group": current_group, "days": md,
                              "display": f"{md} days" if md is not None else "—",
                              "level": level, "is_group": False})
        out["HW Test Plan"] = {"unit": "days", "items": items}

    # ── FW Test Plan ───────────────────────────────────────────────────────
    if "Abbreviation of Test Plan_FW" in wb.sheetnames:
        rows  = list(wb["Abbreviation of Test Plan_FW"].iter_rows(values_only=True))
        items = []
        for i, row in enumerate(rows[1:]):
            name = clean(row[1]) if len(row) > 1 else ""
            if not name: continue
            md = to_float(row[2]) if len(row) > 2 else None
            items.append({"id": f"fw_{i}", "name": name, "group": "FW Tests",
                          "days": md, "display": f"{md} days" if md is not None else "—",
                          "level": None, "is_group": False})
        out["FW Test Plan"] = {"unit": "days", "items": items}

    # ── Module HW ──────────────────────────────────────────────────────────
    if "Abbreviation of TP_Module HW" in wb.sheetnames:
        rows  = list(wb["Abbreviation of TP_Module HW"].iter_rows(values_only=True))
        items = []
        for i, row in enumerate(rows[2:]):
            tc_id = clean(row[0]) if row[0] else f"TC-{i+1}"
            name  = clean(row[1]) if len(row) > 1 else ""
            if not name: continue
            tstr  = clean(row[2]) if len(row) > 2 else ""
            mins  = parse_time_str(tstr)
            days  = round(mins / 480, 3) if mins else None   # 8h × 60min
            display = f"{tstr}" if tstr else "—"
            items.append({"id": f"mhw_{i}", "name": name, "group": tc_id,
                          "days": days, "display": display,
                          "level": None, "is_group": False})
        out["Module HW"] = {"unit": "min", "items": items}

    # ── Comms Testing ──────────────────────────────────────────────────────
    if "Abbreviation of TP_CommsTesting" in wb.sheetnames:
        rows  = list(wb["Abbreviation of TP_CommsTesting"].iter_rows(values_only=True))
        items = []
        for i, row in enumerate(rows[2:]):
            name = clean(row[1]) if len(row) > 1 else ""
            if not name: continue
            hrs  = to_float(row[3]) if len(row) > 3 else None
            days = round(hrs / 8, 2) if hrs is not None else None
            display = f"{hrs} hrs" if hrs is not None else "—"
            items.append({"id": f"ct_{i}", "name": name, "group": "Comms Tests",
                          "days": days, "display": display,
                          "level": None, "is_group": False})
        out["Comms Testing"] = {"unit": "hrs", "items": items}

    wb.close()
    return out

# ─── HTML ─────────────────────────────────────────────────────────────────────
HTML = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>Test Plan Dashboard</title>
<style>
*,*::before,*::after{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Segoe UI',system-ui,sans-serif;background:#0f1117;color:#e2e8f0;min-height:100vh;font-size:13px}

/* ── top bar ── */
.top-bar{background:linear-gradient(135deg,#1e293b,#0f172a);border-bottom:1px solid #334155;
  padding:12px 20px;display:flex;align-items:center;gap:12px}
.logo{width:34px;height:34px;background:linear-gradient(135deg,#6366f1,#8b5cf6);
  border-radius:9px;display:grid;place-items:center;font-size:16px;color:#fff;font-weight:700;flex-shrink:0}
.tb-text{flex:1}
.tb-text h1{font-size:1rem;font-weight:700;color:#f1f5f9;display:flex;align-items:center;gap:7px}
.badge{font-size:.58rem;background:#166534;color:#4ade80;border:1px solid #22543d;padding:2px 7px;border-radius:20px;font-weight:700}
.sub{font-size:.7rem;color:#475569;margin-top:2px}
.tb-actions{display:flex;align-items:center;gap:7px}
.ls{font-size:.67rem;color:#475569}

/* ── tabs ── */
.tabs{display:flex;gap:0;background:#1a2235;border-bottom:2px solid #334155;padding:0 20px}
.tab{padding:9px 18px;font-size:.8rem;font-weight:600;color:#64748b;cursor:pointer;
  border-bottom:2px solid transparent;margin-bottom:-2px;transition:all .2s;white-space:nowrap}
.tab:hover{color:#a5b4fc}
.tab.active{color:#a5b4fc;border-bottom-color:#6366f1}

/* ── buttons ── */
.btn{border-radius:7px;padding:6px 12px;font-size:.77rem;cursor:pointer;
  display:inline-flex;align-items:center;gap:5px;transition:all .2s;
  border:1px solid #334155;background:#1e293b;color:#94a3b8;white-space:nowrap}
.btn:hover{background:#334155;color:#e2e8f0;border-color:#6366f1}
.btn-pur{background:linear-gradient(135deg,#6366f1,#8b5cf6);color:#fff;border-color:transparent;box-shadow:0 2px 8px rgba(99,102,241,.4)}
.btn-pur:hover{opacity:.88;border-color:transparent}
.btn-sm{padding:4px 9px;font-size:.7rem;border-radius:6px}
.btn-danger{border-color:#7f1d1d;color:#f87171}
.btn-danger:hover{background:#7f1d1d;border-color:#f87171;color:#fff}
.btn-success{border-color:#14532d;color:#4ade80}
.btn-success:hover{background:#14532d;border-color:#4ade80;color:#fff}
.spin{animation:spin .7s linear infinite;display:inline-block}
@keyframes spin{to{transform:rotate(360deg)}}

/* ── upload overlay ── */
.overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.78);z-index:200;align-items:center;justify-content:center}
.overlay.show{display:flex}
.upload-box{background:#1e293b;border:2px dashed #6366f1;border-radius:16px;padding:40px 48px;text-align:center;max-width:380px;width:92%;position:relative}
.upload-box.drag{background:#1e1b4b;border-color:#a5b4fc}
.close-btn{position:absolute;top:10px;right:12px;background:none;border:none;color:#475569;font-size:1.3rem;cursor:pointer;padding:2px 7px;border-radius:5px}
.close-btn:hover{color:#e2e8f0;background:#334155}
.prog-bar{height:5px;background:#334155;border-radius:3px;margin-top:14px;overflow:hidden;display:none}
.prog-fill{height:100%;background:linear-gradient(90deg,#6366f1,#8b5cf6);width:0%;transition:width .3s}
.prog-txt{font-size:.75rem;color:#94a3b8;margin-top:6px;min-height:18px}

/* ── DATA layout ── */
.layout{display:flex;height:calc(100vh - 100px)}
.sidebar{width:210px;min-width:210px;background:#1e293b;border-right:1px solid #334155;
  display:flex;flex-direction:column;padding:12px 8px;gap:3px;overflow-y:auto}
.sl{font-size:.6rem;font-weight:700;letter-spacing:.12em;color:#475569;text-transform:uppercase;padding:0 8px 8px}
.sbtn{background:transparent;border:1px solid transparent;border-radius:8px;padding:9px 10px;
  cursor:pointer;text-align:left;color:#94a3b8;font-size:.78rem;font-weight:500;transition:all .18s;line-height:1.4;width:100%}
.sbtn:hover{background:#334155;color:#e2e8f0;border-color:#475569}
.sbtn.active{background:linear-gradient(135deg,#6366f1,#8b5cf6);color:#fff;border-color:transparent;box-shadow:0 3px 10px rgba(99,102,241,.4)}
.cnt{display:inline-block;font-size:.6rem;background:rgba(255,255,255,.18);border-radius:20px;padding:1px 6px;margin-left:4px;vertical-align:middle}

.main{flex:1;display:flex;flex-direction:column;overflow:hidden}
.toolbar{background:#1a2235;border-bottom:1px solid #334155;padding:9px 16px;display:flex;align-items:center;gap:9px;flex-wrap:wrap}
.toolbar h2{font-size:.88rem;font-weight:600;color:#f1f5f9;flex:1}
.chip{background:#334155;border-radius:6px;padding:3px 9px;font-size:.69rem;color:#94a3b8}
.chip span{color:#a5b4fc;font-weight:600}
.sw{position:relative}
.sw input{background:#0f1117;border:1px solid #334155;border-radius:7px;padding:5px 10px 5px 26px;
  color:#e2e8f0;font-size:.78rem;width:180px;outline:none;transition:border-color .2s}
.sw input:focus{border-color:#6366f1}
.si{position:absolute;left:7px;top:50%;transform:translateY(-50%);color:#475569;pointer-events:none;font-size:11px}

.table-wrap{flex:1;overflow:auto;padding:12px 14px}
table{width:100%;border-collapse:collapse}
thead tr{position:sticky;top:0;z-index:10}
thead th{background:#1e293b;border:1px solid #334155;padding:8px 10px;font-weight:600;color:#a5b4fc;
  white-space:nowrap;font-size:.69rem;text-transform:uppercase;letter-spacing:.04em;text-align:left}
tbody tr:nth-child(even){background:#141b2d}
tbody tr:hover{background:#1e3a5f}
tbody td{border:1px solid #1e293b;padding:6px 10px;color:#cbd5e1;max-width:240px;word-break:break-word;vertical-align:top}
.ec{color:#334155;font-style:italic;font-size:.68rem}

/* ── EFFORT layout ── */
.effort-layout{display:flex;height:calc(100vh - 100px);overflow:hidden}
.effort-main{flex:1;display:flex;flex-direction:column;overflow:hidden}

/* effort toolbar */
.etoolbar{background:#1a2235;border-bottom:1px solid #334155;padding:9px 14px;
  display:flex;align-items:center;gap:8px;flex-wrap:wrap}
.etoolbar-title{font-size:.88rem;font-weight:700;color:#f1f5f9;flex:1}
.level-filters{display:flex;gap:4px}
.lf{padding:3px 9px;font-size:.69rem;border-radius:6px;cursor:pointer;
  border:1px solid #334155;background:transparent;color:#64748b;transition:all .18s}
.lf:hover{border-color:#6366f1;color:#a5b4fc}
.lf.on{background:#1e1b4b;border-color:#6366f1;color:#a5b4fc}
.lf.lf1.on{background:#7f1d1d;border-color:#f87171;color:#fca5a5}
.lf.lf2.on{background:#78350f;border-color:#f59e0b;color:#fcd34d}
.lf.lf3.on{background:#14532d;border-color:#22c55e;color:#86efac}

/* effort list */
.effort-scroll{flex:1;overflow-y:auto;padding:12px 14px;display:flex;flex-direction:column;gap:8px}

/* sheet accordion */
.sheet-block{background:#1a2235;border:1px solid #334155;border-radius:10px;overflow:hidden}
.sheet-hdr{display:flex;align-items:center;gap:10px;padding:11px 14px;cursor:pointer;
  background:linear-gradient(135deg,#1e293b,#1a2235);user-select:none;transition:background .15s}
.sheet-hdr:hover{background:#283548}
.sheet-hdr .sarrow{color:#475569;font-size:.8rem;transition:transform .2s;flex-shrink:0}
.sheet-hdr.collapsed .sarrow{transform:rotate(-90deg)}
.sheet-hdr .sname{font-size:.85rem;font-weight:700;color:#f1f5f9;flex:1}
.sheet-hdr .smeta{display:flex;align-items:center;gap:8px}
.sheet-hdr .stotal{font-size:.78rem;font-weight:600;color:#a5b4fc;white-space:nowrap}
.sheet-hdr .scount{font-size:.68rem;color:#64748b;white-space:nowrap}
.sa-wrap{display:flex;align-items:center;gap:5px;flex-shrink:0}
.sa-label{font-size:.7rem;color:#64748b}
.sa-cb{width:14px;height:14px;accent-color:#6366f1;cursor:pointer}

/* group inside sheet */
.group-row{display:flex;align-items:center;gap:8px;padding:7px 14px;
  background:#1e1b4b;border-top:1px solid #312e81;cursor:pointer;user-select:none}
.group-row .garr{font-size:.72rem;color:#6366f1;transition:transform .15s}
.group-row.collapsed .garr{transform:rotate(-90deg)}
.group-row .gname{font-size:.75rem;font-weight:600;color:#a5b4fc;flex:1}
.group-row .gsel{font-size:.67rem;color:#4f46e5}
.group-sa{display:flex;align-items:center;gap:4px;font-size:.67rem;color:#475569}
.group-sa input{width:12px;height:12px;accent-color:#6366f1;cursor:pointer}

/* test row */
.test-list{padding:4px 0}
.trow{display:flex;align-items:center;gap:10px;padding:6px 14px 6px 24px;
  transition:background .12s;cursor:pointer}
.trow:hover{background:#1e293b}
.trow input[type=checkbox]{width:14px;height:14px;accent-color:#6366f1;cursor:pointer;flex-shrink:0}
.trow .tname{flex:1;font-size:.78rem;color:#cbd5e1;line-height:1.35}
.trow .tlabel{font-size:.62rem;padding:1px 6px;border-radius:4px;font-weight:600;flex-shrink:0}
.lv1{background:#7f1d1d;color:#fca5a5}
.lv2{background:#78350f;color:#fcd34d}
.lv3{background:#14532d;color:#86efac}
.trow .ttime{font-size:.75rem;color:#a5b4fc;font-weight:600;flex-shrink:0;min-width:70px;text-align:right}
.trow .ttime.nodata{color:#334155}
.sheet-body{border-top:1px solid #334155}

/* ── summary panel ── */
.summary{width:268px;min-width:268px;background:#1e293b;border-left:1px solid #334155;
  display:flex;flex-direction:column;overflow:hidden}
.sum-top{padding:18px 16px;background:linear-gradient(135deg,#1e1b4b,#1a2235);border-bottom:1px solid #334155;flex-shrink:0}
.sum-label{font-size:.62rem;font-weight:700;letter-spacing:.12em;color:#475569;text-transform:uppercase;margin-bottom:6px}
.sum-big{font-size:2.8rem;font-weight:800;color:#a5b4fc;line-height:1;letter-spacing:-.02em}
.sum-unit{font-size:.72rem;color:#475569;margin-top:3px}
.sum-sub{font-size:.72rem;color:#94a3b8;margin-top:8px}

.sum-body{flex:1;overflow-y:auto;padding:14px 16px;display:flex;flex-direction:column;gap:12px}
.sum-section{}
.sum-section-title{font-size:.62rem;font-weight:700;letter-spacing:.1em;color:#475569;
  text-transform:uppercase;margin-bottom:8px}
.sum-sheet-row{display:flex;align-items:center;gap:6px;padding:6px 8px;
  border-radius:7px;margin-bottom:3px;background:#1a2235;border:1px solid #334155}
.sum-sheet-row .ssname{flex:1;font-size:.73rem;color:#94a3b8;font-weight:500}
.sum-sheet-row .sscount{font-size:.68rem;color:#475569}
.sum-sheet-row .ssval{font-size:.73rem;color:#a5b4fc;font-weight:600;min-width:58px;text-align:right}
.sum-divider{border:none;border-top:1px solid #334155;margin:4px 0}
.sum-stat-row{display:flex;justify-content:space-between;align-items:center;font-size:.75rem;padding:2px 0}
.sum-stat-row .sk{color:#64748b}
.sum-stat-row .sv{color:#e2e8f0;font-weight:600}
.sum-footer{padding:12px 16px;border-top:1px solid #334155;flex-shrink:0;display:flex;flex-direction:column;gap:7px}
.full-btn{width:100%;border-radius:7px;padding:7px;font-size:.76rem;cursor:pointer;
  border:1px solid;text-align:center;transition:all .2s;font-weight:500}
.full-btn.clear{border-color:#7f1d1d;color:#f87171;background:transparent}
.full-btn.clear:hover{background:#7f1d1d;color:#fff}
.full-btn.selall{border-color:#14532d;color:#4ade80;background:transparent}
.full-btn.selall:hover{background:#14532d;color:#fff}

/* state */
.state-box{flex:1;display:flex;flex-direction:column;align-items:center;
  justify-content:center;gap:12px;color:#475569;text-align:center;padding:40px}
.state-box .big{font-size:2.4rem}
.state-box p{font-size:.85rem;line-height:1.6}

::-webkit-scrollbar{width:5px;height:5px}
::-webkit-scrollbar-track{background:#0f1117}
::-webkit-scrollbar-thumb{background:#334155;border-radius:3px}
</style>
</head>
<body>

<!-- Upload overlay -->
<div class="overlay" id="overlay" onclick="overlayClick(event)">
  <div class="upload-box" id="upload-box"
       ondragover="doDragOver(event)" ondragleave="doDragLeave()" ondrop="doDrop(event)">
    <button class="close-btn" onclick="closeUpload()">✕</button>
    <div style="font-size:2rem;margin-bottom:10px">📂</div>
    <h2 style="color:#f1f5f9;font-size:1rem;margin-bottom:6px">Upload Excel File</h2>
    <p style="color:#64748b;font-size:.8rem;margin-bottom:18px">Drag & drop .xlsx or click below</p>
    <button class="btn btn-pur" onclick="document.getElementById('fi').click()">📎 Choose File</button>
    <input type="file" id="fi" style="display:none" accept=".xlsx,.xls" onchange="doUpload(this.files[0])"/>
    <div class="prog-bar" id="prog-bar"><div class="prog-fill" id="prog-fill"></div></div>
    <div class="prog-txt" id="prog-txt"></div>
  </div>
</div>

<!-- Top bar -->
<div class="top-bar">
  <div class="logo">T</div>
  <div class="tb-text">
    <h1>Test Plan Dashboard <span class="badge">LIVE</span></h1>
    <div class="sub" id="sub">Loading…</div>
  </div>
  <div class="tb-actions">
    <span class="ls" id="ls"></span>
    <button class="btn" onclick="loadAll()"><span id="si">🔄</span> Refresh</button>
    <button class="btn btn-pur" onclick="openUpload()">⬆ Upload Excel</button>
  </div>
</div>

<!-- Tabs -->
<div class="tabs">
  <div class="tab active" id="tab-data"   onclick="switchTab('data')">📊 Data View</div>
  <div class="tab"        id="tab-effort" onclick="switchTab('effort')">⏱ Effort Calculator</div>
</div>

<!-- ══════════════════════════ DATA TAB ══════════════════════════ -->
<div class="layout" id="pane-data">
  <nav class="sidebar" id="sidebar"><div class="sl">Sheets</div></nav>
  <div class="main">
    <div class="state-box" id="data-state"><div class="big">⏳</div><p>Loading…</p></div>
    <div id="data-content" style="display:none;flex-direction:column;flex:1;overflow:hidden">
      <div class="toolbar">
        <h2 id="sheet-title"></h2>
        <div style="display:flex;gap:6px">
          <div class="chip">Rows <span id="rc">0</span></div>
          <div class="chip">Cols <span id="cc">0</span></div>
        </div>
        <div class="sw"><span class="si">🔍</span>
          <input type="text" id="q" placeholder="Search…" oninput="filterData()"/></div>
      </div>
      <div class="table-wrap">
        <table><thead id="thead"></thead><tbody id="tbody"></tbody></table>
      </div>
    </div>
  </div>
</div>

<!-- ══════════════════════════ EFFORT TAB ══════════════════════════ -->
<div class="effort-layout" id="pane-effort" style="display:none">
  <div class="effort-main">
    <!-- effort toolbar -->
    <div class="etoolbar">
      <span class="etoolbar-title">⏱ Effort Calculator</span>
      <div class="sw"><span class="si">🔍</span>
        <input type="text" id="eq" placeholder="Search tests…" oninput="renderEffort()"/></div>
      <div class="level-filters">
        <button class="lf on"  id="lf0" onclick="setLvFilter(0)">All</button>
        <button class="lf lf1" id="lf1" onclick="setLvFilter(1)">🔴 Critical</button>
        <button class="lf lf2" id="lf2" onclick="setLvFilter(2)">🟡 Hard</button>
        <button class="lf lf3" id="lf3" onclick="setLvFilter(3)">🟢 Others</button>
      </div>
    </div>
    <div class="state-box" id="effort-state" style="display:none"><div class="big">⏳</div><p>Loading…</p></div>
    <div class="effort-scroll" id="effort-scroll"></div>
  </div>

  <!-- Summary panel -->
  <div class="summary">
    <div class="sum-top">
      <div class="sum-label">Grand Total Effort</div>
      <div class="sum-big" id="grand-total">0</div>
      <div class="sum-unit">Man Days</div>
      <div class="sum-sub" id="grand-sub"></div>
    </div>
    <div class="sum-body">
      <div class="sum-section">
        <div class="sum-section-title">By Sheet</div>
        <div id="sum-sheets"></div>
      </div>
      <hr class="sum-divider"/>
      <div class="sum-section">
        <div class="sum-section-title">Overview</div>
        <div class="sum-stat-row"><span class="sk">Tests selected</span><span class="sv" id="sum-sel">0</span></div>
        <div class="sum-stat-row"><span class="sk">HW Critical</span><span class="sv" id="sum-lv1" style="color:#fca5a5">0 days</span></div>
        <div class="sum-stat-row"><span class="sk">HW Hard to pass</span><span class="sv" id="sum-lv2" style="color:#fcd34d">0 days</span></div>
        <div class="sum-stat-row"><span class="sk">HW Others</span><span class="sv" id="sum-lv3" style="color:#86efac">0 days</span></div>
        <div class="sum-stat-row"><span class="sk">No time data</span><span class="sv" id="sum-nd">0</span></div>
      </div>
    </div>
    <div class="sum-footer">
      <button class="full-btn selall" onclick="selectAllGlobal()">✓ Select All Tests</button>
      <button class="full-btn clear"  onclick="clearAllGlobal()">✕ Clear All</button>
    </div>
  </div>
</div>

<script>
/* ═══════════════════════════ state ═══════════════════════════ */
let DATA={}, ROWS=[], ACTIVE_SHEET=null;
let EFFORT={};       // { sheetName: { unit, items[] } }
let SEL=new Set();   // selected test ids
let LV_FILTER=0;     // 0=all 1=critical 2=hard 3=others
let collapsed={};    // id -> bool

const SHEET_ICONS={"Priority Test":"⭐","HW Test Plan":"🔧","FW Test Plan":"💾","Module HW":"📡","Comms Testing":"📶"};
const SHEET_UNIT_LABEL={"Priority Test":"days","HW Test Plan":"days","FW Test Plan":"days","Module HW":"min","Comms Testing":"hrs"};

/* ═══════════════════════════ tabs ═══════════════════════════ */
function switchTab(t){
  document.getElementById('pane-data').style.display   = t==='data'  ?'flex':'none';
  document.getElementById('pane-effort').style.display = t==='effort'?'flex':'none';
  ['data','effort'].forEach(x=>document.getElementById('tab-'+x).classList.toggle('active',x===t));
}

/* ═══════════════════════════ load ═══════════════════════════ */
async function fetchWithRetry(url,retries=5,delay=6000){
  for(let i=0;i<retries;i++){
    try{
      const ctrl=new AbortController();
      const tid=setTimeout(()=>ctrl.abort(),55000);
      const r=await fetch(url,{signal:ctrl.signal});
      clearTimeout(tid);
      if(r.ok) return r;
      throw new Error('HTTP '+r.status);
    }catch(e){
      if(i===retries-1) throw e;
      const w=delay*(i+1);
      setDataMsg(`Server waking up… retry ${i+1}/${retries} in ${w/1000}s`);
      await new Promise(res=>setTimeout(res,w));
    }
  }
}

function setDataMsg(m){
  document.getElementById('data-state').innerHTML='<div class="big">⏳</div><p style="color:#a5b4fc">'+m+'</p>';
  document.getElementById('data-state').style.display='flex';
  document.getElementById('data-content').style.display='none';
}

async function loadAll(){
  document.getElementById('si').classList.add('spin');
  await Promise.all([loadData(),loadEffort()]);
  document.getElementById('si').classList.remove('spin');
}

async function loadData(){
  try{
    setDataMsg('Connecting…');
    const r=await fetchWithRetry('/data?t='+Date.now());
    const j=await r.json();
    DATA=j.data;
    document.getElementById('sub').textContent=j.filename+' · Real-time from Excel';
    document.getElementById('ls').textContent='Synced '+new Date().toLocaleTimeString();
    buildSidebar();
    if(ACTIVE_SHEET&&DATA[ACTIVE_SHEET]) showSheet(ACTIVE_SHEET);
    else if(Object.keys(DATA).length){ACTIVE_SHEET=Object.keys(DATA)[0];buildSidebar();showSheet(ACTIVE_SHEET);}
  }catch(e){
    document.getElementById('data-state').innerHTML=
      '<div class="big">❌</div><p style="color:#f87171">'+e.message+'</p>'+
      '<button class="btn btn-pur" onclick="loadAll()" style="margin-top:8px">🔄 Retry</button>';
    document.getElementById('data-state').style.display='flex';
    document.getElementById('data-content').style.display='none';
  }
}

async function loadEffort(){
  try{
    const r=await fetchWithRetry('/effort?t='+Date.now());
    EFFORT=await r.json();
    renderEffort();
    document.getElementById('effort-state').style.display='none';
  }catch(e){
    document.getElementById('effort-state').innerHTML=
      '<div class="big">❌</div><p style="color:#f87171">'+e.message+'</p>';
    document.getElementById('effort-state').style.display='flex';
  }
}

/* ═══════════════════════════ DATA tab ═══════════════════════════ */
function buildSidebar(){
  const sb=document.getElementById('sidebar');
  sb.innerHTML='<div class="sl">Sheets</div>';
  Object.keys(DATA).forEach(s=>{
    const b=document.createElement('button');
    b.className='sbtn'+(s===ACTIVE_SHEET?' active':'');
    b.innerHTML=(SHEET_ICONS[s]||'📄')+' '+s+'<span class="cnt">'+DATA[s].rows.length+'</span>';
    b.onclick=()=>{ACTIVE_SHEET=s;buildSidebar();showSheet(s);};
    sb.appendChild(b);
  });
}

function showSheet(s){
  const{headers,rows:r}=DATA[s]; ROWS=r;
  document.getElementById('sheet-title').textContent=s;
  document.getElementById('rc').textContent=r.length;
  document.getElementById('cc').textContent=headers.length;
  document.getElementById('q').value='';
  document.getElementById('thead').innerHTML='<tr>'+headers.map(h=>'<th>'+esc(h)+'</th>').join('')+'</tr>';
  renderRows(r);
  document.getElementById('data-state').style.display='none';
  document.getElementById('data-content').style.display='flex';
}

function renderRows(data){
  document.getElementById('tbody').innerHTML=data.map(r=>
    '<tr>'+r.map(c=>c?'<td>'+esc(c)+'</td>':'<td><span class="ec">—</span></td>').join('')+'</tr>'
  ).join('');
  document.getElementById('rc').textContent=data.length;
}

function filterData(){
  const q=document.getElementById('q').value.toLowerCase().trim();
  renderRows(q?ROWS.filter(r=>r.some(c=>c.toLowerCase().includes(q))):ROWS);
}

/* ═══════════════════════════ EFFORT tab ═══════════════════════════ */
function setLvFilter(lv){
  LV_FILTER=lv;
  [0,1,2,3].forEach(i=>document.getElementById('lf'+i).classList.toggle('on',i===lv));
  renderEffort();
}

function renderEffort(){
  const q=(document.getElementById('eq').value||'').toLowerCase().trim();
  const scroll=document.getElementById('effort-scroll');
  scroll.innerHTML='';

  Object.entries(EFFORT).forEach(([sheetName,{unit,items}])=>{
    const testItems=items.filter(it=>!it.is_group);
    // Apply filters
    const visible=testItems.filter(it=>{
      if(LV_FILTER>0 && sheetName==='HW Test Plan' && it.level!==LV_FILTER) return false;
      if(q && !it.name.toLowerCase().includes(q) && !it.group.toLowerCase().includes(q)) return false;
      return true;
    });
    if(!visible.length && q) return; // hide empty sheets when searching

    const selInSheet=visible.filter(it=>SEL.has(it.id));
    const totalDays=selInSheet.reduce((s,t)=>s+(t.days||0),0);

    // Sheet block
    const block=document.createElement('div');
    block.className='sheet-block';
    const isCollapsed=collapsed['sheet_'+sheetName]||false;

    // Sheet header
    const hdr=document.createElement('div');
    hdr.className='sheet-hdr'+(isCollapsed?' collapsed':'');
    hdr.innerHTML=`
      <span class="sarrow">▾</span>
      <span class="sname">${SHEET_ICONS[sheetName]||'📄'} ${sheetName}</span>
      <div class="smeta">
        <span class="scount">${selInSheet.length}/${visible.length} selected</span>
        <span class="stotal">${totalDays>0?totalDays.toFixed(1)+' days':''}</span>
      </div>
      <div class="sa-wrap" onclick="event.stopPropagation()">
        <span class="sa-label">All</span>
        <input type="checkbox" class="sa-cb"
          ${selInSheet.length===visible.length&&visible.length>0?'checked':''}
          onchange="toggleSheet('${sheetName}',this.checked)"/>
      </div>`;
    hdr.addEventListener('click',()=>{ collapsed['sheet_'+sheetName]=!collapsed['sheet_'+sheetName]; renderEffort(); });
    block.appendChild(hdr);

    if(!isCollapsed){
      const body=document.createElement('div');
      body.className='sheet-body';

      // Group items or flat list
      if(sheetName==='HW Test Plan'){
        // Grouped by section
        const groups={};
        const groupOrder=[];
        items.filter(it=>it.is_group).forEach(g=>{groups[g.name]=[]; groupOrder.push(g.name);});
        visible.forEach(it=>{ if(groups[it.group]) groups[it.group].push(it); });
        groupOrder.forEach(grpName=>{
          const grpTests=groups[grpName];
          if(!grpTests||!grpTests.length) return;
          const grpCollapsed=collapsed['grp_'+grpName]||false;
          const grpSel=grpTests.filter(t=>SEL.has(t.id)).length;
          const grpRow=document.createElement('div');
          grpRow.className='group-row'+(grpCollapsed?' collapsed':'');
          grpRow.innerHTML=`
            <span class="garr">▾</span>
            <span class="gname">${esc(grpName)}</span>
            <span class="gsel">${grpSel}/${grpTests.length}</span>
            <div class="group-sa" onclick="event.stopPropagation()">
              <input type="checkbox" ${grpSel===grpTests.length&&grpTests.length>0?'checked':''}
                onchange="toggleGroup('${grpName.replace(/'/g,"\\'")}',this.checked)"/>
              <span>All</span>
            </div>`;
          grpRow.addEventListener('click',()=>{collapsed['grp_'+grpName]=!collapsed['grp_'+grpName];renderEffort();});
          body.appendChild(grpRow);
          if(!grpCollapsed){
            const list=document.createElement('div');
            list.className='test-list';
            grpTests.forEach(it=>list.appendChild(makeTestRow(it,'HW Test Plan')));
            body.appendChild(list);
          }
        });
      } else {
        const list=document.createElement('div');
        list.className='test-list';
        visible.forEach(it=>list.appendChild(makeTestRow(it,sheetName)));
        body.appendChild(list);
      }
      block.appendChild(body);
    }
    scroll.appendChild(block);
  });

  updateSummary();
}

function makeTestRow(it, sheetName){
  const row=document.createElement('div');
  row.className='trow';
  const lvLabel=it.level===1?'Critical':it.level===2?'Hard':it.level===3?'Others':'';
  const lvClass=it.level===1?'lv1':it.level===2?'lv2':it.level===3?'lv3':'';
  const hasTime=it.days!=null;
  row.innerHTML=
    `<input type="checkbox" ${SEL.has(it.id)?'checked':''} onchange="toggleTest('${it.id}')"/>
     <span class="tname">${esc(it.name)}</span>
     ${lvLabel?`<span class="tlabel ${lvClass}">${lvLabel}</span>`:''}
     <span class="ttime ${hasTime?'':'nodata'}">${it.display}</span>`;
  row.querySelector('input').addEventListener('change',function(){ toggleTest(it.id,this.checked); });
  row.addEventListener('click',e=>{ if(e.target.tagName!=='INPUT'){ const cb=row.querySelector('input'); cb.checked=!cb.checked; toggleTest(it.id,cb.checked); }});
  return row;
}

function toggleTest(id,force){
  if(force===undefined) force=!SEL.has(id);
  if(force) SEL.add(id); else SEL.delete(id);
  renderEffort();
}

function toggleSheet(sheetName,on){
  const {items}=EFFORT[sheetName];
  const q=(document.getElementById('eq').value||'').toLowerCase().trim();
  items.filter(it=>!it.is_group).forEach(it=>{
    if(LV_FILTER>0 && sheetName==='HW Test Plan' && it.level!==LV_FILTER) return;
    if(q && !it.name.toLowerCase().includes(q)) return;
    if(on) SEL.add(it.id); else SEL.delete(it.id);
  });
  renderEffort();
}

function toggleGroup(grpName,on){
  Object.values(EFFORT).forEach(({items})=>{
    items.filter(it=>!it.is_group&&it.group===grpName).forEach(it=>{
      if(on) SEL.add(it.id); else SEL.delete(it.id);
    });
  });
  renderEffort();
}

function selectAllGlobal(){
  Object.values(EFFORT).forEach(({items})=>items.filter(it=>!it.is_group).forEach(it=>SEL.add(it.id)));
  renderEffort();
}

function clearAllGlobal(){ SEL.clear(); renderEffort(); }

function updateSummary(){
  let grandDays=0, totalSel=0, noData=0;
  let lv1=0,lv2=0,lv3=0;
  const sheetRows=[];

  Object.entries(EFFORT).forEach(([name,{unit,items}])=>{
    const tests=items.filter(it=>!it.is_group);
    const sel=tests.filter(it=>SEL.has(it.id));
    const days=sel.reduce((s,t)=>s+(t.days||0),0);
    grandDays+=days;
    totalSel+=sel.length;
    noData+=sel.filter(t=>t.days==null).length;
    if(name==='HW Test Plan'){
      lv1+=sel.filter(t=>t.level===1).reduce((s,t)=>s+(t.days||0),0);
      lv2+=sel.filter(t=>t.level===2).reduce((s,t)=>s+(t.days||0),0);
      lv3+=sel.filter(t=>t.level===3).reduce((s,t)=>s+(t.days||0),0);
    }
    sheetRows.push({name,sel:sel.length,total:tests.length,days,unit});
  });

  document.getElementById('grand-total').textContent=grandDays.toFixed(1);
  document.getElementById('grand-sub').textContent=totalSel+' test'+(totalSel!==1?'s':'')+' selected';
  document.getElementById('sum-sel').textContent=totalSel;
  document.getElementById('sum-lv1').textContent=lv1.toFixed(1)+' days';
  document.getElementById('sum-lv2').textContent=lv2.toFixed(1)+' days';
  document.getElementById('sum-lv3').textContent=lv3.toFixed(1)+' days';
  document.getElementById('sum-nd').textContent=noData;

  const sumSheets=document.getElementById('sum-sheets');
  sumSheets.innerHTML=sheetRows.map(r=>`
    <div class="sum-sheet-row">
      <span class="ssname">${SHEET_ICONS[r.name]||'📄'} ${r.name}</span>
      <span class="sscount">${r.sel}/${r.total}</span>
      <span class="ssval">${r.days>0?r.days.toFixed(1)+'d':'—'}</span>
    </div>`).join('');
}

/* ═══════════════════════════ upload ═══════════════════════════ */
function openUpload(){ document.getElementById('overlay').classList.add('show'); resetUpload(); }
function closeUpload(){ document.getElementById('overlay').classList.remove('show'); }
function overlayClick(e){ if(e.target===document.getElementById('overlay')) closeUpload(); }
function doDragOver(e){ e.preventDefault(); document.getElementById('upload-box').classList.add('drag'); }
function doDragLeave(){ document.getElementById('upload-box').classList.remove('drag'); }
function doDrop(e){ e.preventDefault(); doDragLeave(); if(e.dataTransfer.files[0]) doUpload(e.dataTransfer.files[0]); }

async function doUpload(file){
  if(!file||!file.name.match(/\.xlsx?$/i)){ alert('Please select .xlsx or .xls file'); return; }
  const bar=document.getElementById('prog-bar'),fill=document.getElementById('prog-fill'),txt=document.getElementById('prog-txt');
  bar.style.display='block'; fill.style.width='30%'; txt.textContent='Uploading '+file.name+'…';
  const fd=new FormData(); fd.append('file',file);
  try{
    fill.style.width='65%';
    const r=await fetch('/upload',{method:'POST',body:fd});
    fill.style.width='90%';
    if(!r.ok) throw new Error(await r.text());
    fill.style.width='100%';
    txt.innerHTML='<span style="color:#4ade80">✓ Done! Refreshing…</span>';
    setTimeout(async()=>{ await loadAll(); closeUpload(); },700);
  }catch(e){
    txt.innerHTML='<span style="color:#f87171">✗ '+e.message+'</span>';
    fill.style.width='0%';
  }
}

function resetUpload(){
  document.getElementById('prog-bar').style.display='none';
  document.getElementById('prog-fill').style.width='0%';
  document.getElementById('prog-txt').textContent='';
  document.getElementById('fi').value='';
}

/* ═══════════════════════════ helpers ═══════════════════════════ */
function esc(s){ return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }

/* ═══════════════════════════ init ═══════════════════════════ */
loadAll();
setInterval(loadAll,30000);
setInterval(()=>fetch('/ping').catch(()=>{}), 10*60*1000);
</script>
</body>
</html>"""

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/data")
def data():
    try:
        return jsonify({"filename": active_xl().name, "data": read_excel()})
    except Exception as e:
        return str(e), 500

@app.route("/effort")
def effort():
    try:
        return jsonify(read_effort())
    except Exception as e:
        return str(e), 500

@app.route("/ping")
def ping():
    return "ok", 200

@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return "No file provided", 400
    f = request.files["file"]
    if not f.filename:
        return "Empty filename", 400
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return "Only .xlsx / .xls files supported", 400
    f.save(str(UPLOADED_XL))
    try:
        wb = openpyxl.load_workbook(UPLOADED_XL, read_only=True)
        wb.close()
    except Exception:
        UPLOADED_XL.unlink(missing_ok=True)
        return "Invalid or corrupt Excel file", 400
    return jsonify({"ok": True, "filename": f.filename})

if __name__ == "__main__":
    app.run(debug=True, port=8765)
